import openpyxl
import openpyxl.styles.alignment

def push():
    book=openpyxl.Workbook()
    sheet=book.active

    def genExcel():
        taxSum=0
        totSum=0
        for i in range(4,sheet.max_row+1):
            taxSum+=int(sheet['O{}'.format(i)].value)
            totSum+=int(sheet['P{}'.format(i)].value)
        sheet['O1']=taxSum
        sheet['O2']=totSum

    header1=["ANTARTICA COLDWEARS","","","","PURCHASE","","",month,"","",year,"","","Total Tax","",""]
    header2=["","","","","","","","","","","","","","Total Purchase","",""]
    heading=["Invoice No","Invoice Date","GSTIN","Seller's Name","Address","Product Name","HSN Code","Gross Value","IGST","","CGST","","SGST","","Total Tax","Total Invoice Value"]
    sheet.append(header1)
    sheet.append(header2)
    sheet.append(heading)

    sheet.merge_cells('A1:D2')
    sheet.merge_cells('E1:G2')
    sheet.merge_cells('H1:J2')
    sheet.merge_cells('K1:M2')

    sheet.merge_cells('O1:P1')
    sheet.merge_cells('O2:P2')


    sheet.merge_cells('I3:J3')
    sheet.merge_cells('K3:L3')
    sheet.merge_cells('M3:N3')

    sheet.row_dimensions[1].height=20
    sheet.row_dimensions[2].height=20

    sheet.column_dimensions['A'].width=15
    sheet.column_dimensions['B'].width=18
    sheet.column_dimensions['C'].width=23
    sheet.column_dimensions['D'].width=40
    sheet.column_dimensions['E'].width=20
    sheet.column_dimensions['F'].width=25
    sheet.column_dimensions['G'].width=18
    sheet.column_dimensions['H'].width=20
    sheet.column_dimensions['I'].width=10
    sheet.column_dimensions['J'].width=15
    sheet.column_dimensions['K'].width=10
    sheet.column_dimensions['L'].width=15
    sheet.column_dimensions['M'].width=10
    sheet.column_dimensions['N'].width=15
    sheet.column_dimensions['O'].width=17
    sheet.column_dimensions['P'].width=20




    sample1=["","","","","","","","","IGST","","9%","","9%","","400","45000"]
    sample2=["","","","","","","","","IGST","","9%","","9%","","600","65000"]
    sample3=["","","","","","","","","IGST","","9%","","9%","","470","100"]
    sheet.append(sample1)
    sheet.append(sample2)
    sheet.append(sample3)

    genExcel()
    book.save(path)
