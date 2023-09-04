import tkinter
from tkinter import ttk
from tkinter import *
import sv_ttk
from ctypes import windll
from datetime import datetime,date,timedelta
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Alignment
import os,sys
import openpyxl
from tkinter import messagebox
import ntkutils
import webbrowser
from tkinter.simpledialog import askstring
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font,PatternFill,Border,Side
windll.shcore.SetProcessDpiAwareness(1)
basedir = os.path.dirname(__file__)
#_________checker_________________________________________________________________________
def check(gstin):
    import requests
    from bs4 import BeautifulSoup
    global url,inputName,csrfName
    url="https://www.knowyourgst.com/gst-number-search/"
    inputName="gstnum"
    csrfName="csrfmiddlewaretoken"
    client=requests.session()
    global name
    name=""
    client.get(url)
    if(client.get(url).status_code!=200):
        name+="[GET URL Failure]"
    if 'csrftoken' in client.cookies:
        csrftoken = client.cookies['csrftoken']
    else:
        csrftoken = client.cookies['csrf']
    data={inputName:gstin,csrfName:csrftoken}
    r = client.post(url,data=data,headers=dict(Referer=url))
    if (r.content):
        pass
    else:
        name+="[Content retrival failure]"
    try:
        soup=BeautifulSoup(r.text,'html.parser')
        table=soup.find('table',class_="striped highlight questionlist").find_all('tr')
        data=[]
        for tr in table:
            data.append([td.text for td in tr.find_all('td')])
        reqData=[]
        for i in range(len(data)):
                tempList=data[i]
                reqData.append(tempList[1])
        global add
        name = str(reqData[0])
        add=str(reqData[6].split(",")[0])
    except:
        name+="[Parsing Failure]"
#________________________________________________________________________________________
#excel function

def openBook(path):
    book=openpyxl.Workbook()
    sheet=book.active
    f=open(resource_path("assets\\hl290sk"),"r")
    title=f.read()
    header1=[title,"","","","PURCHASE","","",entryMonthEntry.get().upper(),"","",entryYearEntry.get(),"","","Total Tax","",""]
    header2=["","","","","","","","","","","","","","Total Purchase","",""]
    heading=["Invoice No","Invoice Date","GSTIN","Seller's Name","Address","Product Name","HSN Code","Gross Value","IGST","","CGST","","SGST","","Total Tax","Total Invoice Value"]
    sheet.append(header1)
    sheet.append(header2)
    sheet.append(heading)

    sheet['A1'].border=Border(bottom=Side(border_style='thin',color='000000'))
    sheet['E1'].border=Border(bottom=Side(border_style='thin',color='000000'))
    sheet['H1'].border=Border(bottom=Side(border_style='thin',color='000000'))
    sheet['K1'].border=Border(bottom=Side(border_style='thin',color='000000'))

    sheet.merge_cells('A1:D2')
    sheet.merge_cells('E1:G2')
    sheet.merge_cells('H1:J2')
    sheet.merge_cells('K1:M2')

    sheet.merge_cells('O1:P1')
    sheet.merge_cells('O2:P2')


    sheet.merge_cells('I3:J3')
    sheet.merge_cells('K3:L3')
    sheet.merge_cells('M3:N3')

    sheet.row_dimensions[1].height=20
    sheet.row_dimensions[2].height=20

    sheet.column_dimensions['A'].width=15
    sheet.column_dimensions['B'].width=18
    sheet.column_dimensions['C'].width=23
    sheet.column_dimensions['D'].width=40
    sheet.column_dimensions['E'].width=20
    sheet.column_dimensions['F'].width=25
    sheet.column_dimensions['G'].width=18
    sheet.column_dimensions['H'].width=20
    sheet.column_dimensions['I'].width=10
    sheet.column_dimensions['J'].width=15
    sheet.column_dimensions['K'].width=10
    sheet.column_dimensions['L'].width=15
    sheet.column_dimensions['M'].width=10
    sheet.column_dimensions['N'].width=15
    sheet.column_dimensions['O'].width=17
    sheet.column_dimensions['P'].width=22

    sheet['A1'].alignment=Alignment(horizontal="center",vertical="center")
    sheet['E1'].alignment=Alignment(horizontal="center",vertical="center")
    sheet['H1'].alignment=Alignment(horizontal="center",vertical="center")
    sheet['K1'].alignment=Alignment(horizontal="center",vertical="center")

    sheet['I3'].alignment=Alignment(horizontal="center")
    sheet['K3'].alignment=Alignment(horizontal="center")
    sheet['M3'].alignment=Alignment(horizontal="center")

    sheet['A1'].font=Font(name="SimSun",size=36,bold=True)
    sheet['E1'].font=Font(name="SimSun",size=36)
    sheet['H1'].font=Font(name="SimSun",size=36)
    sheet['K1'].font=Font(name="SimSun",size=36)

    sheet['A1'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')
    sheet['E1'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')
    sheet['H1'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')
    sheet['K1'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')

    sheet['N1'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['N2'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['O1'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')
    sheet['O2'].fill=PatternFill(fill_type='solid',start_color='E7E6E6',end_color='E7E6E6')

    sheet['N1'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    sheet['N2'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    sheet['O1'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    sheet['O2'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    sheet['P1'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    sheet['P2'].border=Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))

    sheet['A3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['B3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['C3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['D3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['E3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['F3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['G3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['H3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['I3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['J3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['K3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['L3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['M3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['N3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['O3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')
    sheet['P3'].fill=PatternFill(fill_type='solid',start_color='D0CECE',end_color='D0CECE')

    sheet['A3'].font=Font(name="SimSun")
    sheet['B3'].font=Font(name="SimSun")
    sheet['C3'].font=Font(name="SimSun")
    sheet['D3'].font=Font(name="SimSun")
    sheet['E3'].font=Font(name="SimSun")
    sheet['F3'].font=Font(name="SimSun")
    sheet['G3'].font=Font(name="SimSun")
    sheet['H3'].font=Font(name="SimSun")
    sheet['I3'].font=Font(name="SimSun")
    sheet['J3'].font=Font(name="SimSun")
    sheet['K3'].font=Font(name="SimSun")
    sheet['L3'].font=Font(name="SimSun")
    sheet['M3'].font=Font(name="SimSun")
    sheet['N3'].font=Font(name="SimSun")
    sheet['O3'].font=Font(name="SimSun")
    sheet['P3'].font=Font(name="SimSun")
    sheet['O1']="=SUM(O4:O100)"
    sheet['O2']="=SUM(P4:P100)"
    book.save(path)


#_______________________________________________________________________________________

root=tkinter.Tk()
#f=open("assets\\ejk9304h","r")
themeSet="dark"
if themeSet=="dark":
    ntkutils.dark_title_bar(root)
root.title("fullsafe")
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
tk_title="fullsafe"
root.iconbitmap(default=resource_path("assets\\logo.ico"))

def openFile():
    month=entryMonthEntry.get()
    year=entryYearEntry.get()
    bookpath=month+"_"+year
    path="D:\\fullsafe\\"+year+"\\"+month+"\\"
    Path=path+"Purchase_"+bookpath+".xlsx"
    if os.path.exists(Path):
        os.startfile(Path)
    else:
        messagebox.showerror('File Error','Sumbit atleast one entry for the file to be created')
def optionsName():
    #args=['assets\hl290sk']
    #f=subprocess.call(args,shell=True)
    nme=askstring('Set Business Name','Enter your business name')
    f=open("assets\\hl290sk","w")
    f.write(nme)
    month=entryMonthEntry.get()
    year=entryYearEntry.get()
    bookpath=month+"_"+year
    path="D:\\fullsafe\\"+year+"\\"+month+"\\"
    Path=path+"Purchase_"+bookpath+".xlsx"
    if os.path.exists(Path):
        wb=openpyxl.load_workbook(Path)
        sh=wb.active
        f=open(resource_path("assets\\hl290sk"),"r")
        sh['A1']=f.read()
        wb.save(Path)
def theme():
    global themeSet
    ask=messagebox.askquestion('Application restart','All data entered now in the fields will be lost. Data sumbitted will not be lost.')
    if ask=="yes":
        f=open("assets\\ejk9304h","w")
        if themeSet=="dark":
            f.write("light")
            f=open("assets\\ejk9304h","r")
            themeSet=f.read()
            python = sys.executable
            os.execl(python, python, * sys.argv)
        else:
            f.write("dark")
            f=open("assets\\ejk9304h","r")
            themeSet=f.read()
            python = sys.executable
            os.execl(python, python, * sys.argv)
        sv_ttk.set_theme(themeSet)
    else:
        pass
s=ttk.Style()
s.configure('bbb',font=('Adobe Heiti Std R',20))
def chang():
    f=open("assets\\os20402sk","r")
    if (f.read()=="larg"):
        list1=[entryMonth,entryYear,invoiceNo,invoiceDate,gstNo,bName,bAdd,product,hsn,sub]
        for i in list1:
            i.config(font=("Adobe Heiti Std R",13))
        list2=[extra,taxLabel,totalLabel]
        for i in list2:
            i.config(font=("Adobe Heiti Std R",9))
        list3=[entryMonthEntry,entryYearEntry,invoiceNoEntry,invoiceDateEntry,gstNoEntry,bNameEntry,bAddEntry,productEntry,hsnEntry,subEntry,extraEntry,taxEntry,totalEntry]
        for i in list3:
            i.config(font=(0,10))
    else:
        f=open("assets\\os20402sk","w")
        f.write("small")
def fontChange():
    f=open("assets\\os20402sk","r")
    if (f.read()=="small"):
        list1=[entryMonth,entryYear,invoiceNo,invoiceDate,gstNo,bName,bAdd,product,hsn,sub]
        for i in list1:
            i.config(font=("Adobe Heiti Std R",13))
        list2=[extra,taxLabel,totalLabel]
        for i in list2:
            i.config(font=("Adobe Heiti Std R",9))
        list3=[invoiceDateEntry,entryMonthEntry,entryYearEntry,invoiceNoEntry,gstNoEntry,bNameEntry,bAddEntry,productEntry,hsnEntry,subEntry,extraEntry,taxEntry,totalEntry]
        for i in list3:
            i.config(font=(0,10))
        g=open("assets\\os20402sk","w")
        g.write=("larg")
    else:
        python = sys.executable
        os.execl(python, python, * sys.argv)
optionFrame=ttk.Frame(root)
optionFrame.pack(anchor=W,side=tkinter.TOP)
lab1=tkinter.Button(optionFrame,text="File",border=0,fg='#adb5bd',command=openFile).pack(side=tkinter.LEFT,padx=(5,7),pady=(0,6))
lab2=tkinter.Button(optionFrame,text="Options",border=0,fg='#adb5bd',command=optionsName).pack(side=tkinter.LEFT,padx=(0,7),pady=(0,6))
lab3=tkinter.Button(optionFrame,text="Font",border=0,fg='#adb5bd',command=fontChange).pack(side=tkinter.LEFT,pady=(0,6),padx=(0,7))
lab4=tkinter.Button(optionFrame,text="View",border=0,fg='#adb5bd',command=theme).pack(side=tkinter.LEFT,pady=(0,6),padx=(0,7))
lab5=tkinter.Button(optionFrame,text="Help",border=0,fg='#adb5bd',command=lambda : webbrowser.open("https://github.com/hariprasath112/fullsafe",new=0,autoraise=True)).pack(side=tkinter.LEFT,pady=(0,6),padx=(0,7))

parent=ttk.Frame(root)
parent.pack(side=tkinter.TOP,anchor=N)
frame=ttk.Frame(parent)
frame.grid(row=0,column=0,sticky=N)
#entry config frame
entryDateFrame=ttk.LabelFrame(frame,text="Entry Configuration",)
entryDateFrame.grid(row=0,column=0, sticky=N)
#font=(None,12)
#___________________________________________________________________________________________________________________________
#label for first row entry date and month
entryMonth=ttk.Label(entryDateFrame,text="Entry Month")
entryMonth.grid(row=0,column=0,padx=30,pady=10)
entryYear=ttk.Label(entryDateFrame,text="Entry Year")
entryYear.grid(row=0,column=1,padx=30,pady=10)
#input boxes for entry date and month - second row
entryMonthEntry=ttk.Combobox(entryDateFrame,state="readonly",values=["January","Febuary","March","April","May","June","July","August","September","October","Novemeber","December"])
entryMonthEntry.grid(row=1,column=0,padx=30,pady=10)
entryMonthEntry.current((int(datetime.now().strftime('%m'))-2))
entryYearEntry=ttk.Entry(entryDateFrame)
entryYearEntry.insert(0,(date.today().replace(day=1) - timedelta(days=1)).year)
entryYearEntry.grid(row=1,column=1,padx=30,pady=10)
#___________________________________________________________________________________________________________________________
#invoice frame
invoiceFrame=ttk.LabelFrame(frame,text="Invoice Details")
invoiceFrame.grid(row=1,column=0, sticky=E+W)
#invoiceNO details
invoiceNo=ttk.Label(invoiceFrame,text="Invoice No.")
invoiceNo.grid(row=0,column=0,padx=30,pady=10, sticky="ew")
invoiceNoEntry=ttk.Entry(invoiceFrame)
invoiceNoEntry.grid(row=0,column=1,padx=30,pady=10, sticky="ew")
invoiceNoEntry.focus_set()
#invoiceDate
invoiceDate=ttk.Label(invoiceFrame,text="Invoice Date")
invoiceDate.grid(row=1,column=0,padx=30,pady=10, sticky="ew")
invoiceDateEntry=DateEntry(invoiceFrame,showothermonthdays=False,showweeknumbers=False,date_pattern='dd/mm/yyyy',month=(int(datetime.now().strftime('%m'))-1),year=(date.today().replace(day=1) - timedelta(days=1)).year)
invoiceDateEntry.grid(row=1,column=1,padx=30,pady=10, sticky="ew")
invoiceNoEntry.bind('<Return>',lambda e: invoiceDateEntry.focus_set())
#___________________________________________________________________________________________________________________________
#gstin
gstNo=ttk.Label(invoiceFrame,text="GST No.")
gstNo.grid(row=2,column=0,padx=30,pady=10, sticky="ew")
gstNoEntry=ttk.Entry(invoiceFrame)
gstNoEntry.grid(row=2,column=1,padx=30,pady=10, sticky="ew")
invoiceDateEntry.bind('<Return>',lambda e: gstNoEntry.focus_set())
#---------------------------------------------------
gstName=tkinter.StringVar()
gstAddress=tkinter.StringVar()
def retreiveGstNo(event=None):
    global gstNum,gstAddress
    gstNumber=gstNoEntry.get()
    check(gstNumber)
    gstName.set(name)   #for bName
    gstAddress.set(add)    #for bAdd

#button that onclick call func
gstGetButton=ttk.Button(invoiceFrame,command=retreiveGstNo,text="✓")
gstGetButton.grid(row=2,column=2, sticky="e")
def lambdaCall():
    retreiveGstNo()
    productEntry.focus_set()
gstNoEntry.bind('<Return>',lambda e: lambdaCall())
#gstGetButton.bind('<Return>',gstGetButton.invoke())   gstGetButton.focus_set() and 
#-------------------------------------------------------
bName=ttk.Label(invoiceFrame,text="Business Name")
bName.grid(row=3,column=0,padx=30,pady=10, sticky="ew")
bNameEntry=ttk.Entry(invoiceFrame,textvariable=gstName)
bNameEntry.grid(row=3,column=1,padx=30,pady=10, sticky="ew")

bAdd=ttk.Label(invoiceFrame,text="Business Address")
bAdd.grid(row=4,column=0,padx=30,pady=10, sticky="ew")
bAddEntry=ttk.Entry(invoiceFrame,textvariable=gstAddress)
bAddEntry.grid(row=4,column=1,padx=30,pady=10, sticky="ew")
#___________________________________________________________________________________________________________________________
#product frame
productFrame=ttk.LabelFrame(frame,text="Product Details")
productFrame.grid(row=2,column=0, sticky=N)
#productbox
product=ttk.Label(productFrame,text="Product Name")
product.grid(row=0,column=0,padx=30,pady=10, sticky="ew")
productEntry=ttk.Entry(productFrame)
productEntry.grid(row=0,column=1,padx=30,pady=10, sticky="w")
gstGetButton.bind('<Return>',lambda e: productEntry.focus_set())

#hsn
hsn=ttk.Label(productFrame,text="HSN")
hsn.grid(row=1,column=0,padx=30,pady=10, sticky="ew")
hsnEntry=ttk.Entry(productFrame)
hsnEntry.grid(row=1,column=1,padx=30,pady=10, sticky="w")
productEntry.bind('<Return>',lambda e: hsnEntry.focus_set())

#subtotal
sub=ttk.Label(productFrame,text="Sub Total")
sub.grid(row=2,column=0,padx=30,pady=10, sticky="w")
subEntry=ttk.Entry(productFrame)
subEntry.grid(row=2,column=1,padx=30,pady=10, sticky="w")
hsnEntry.bind('<Return>',lambda e: subEntry.focus_set())


#extra
extraFrame=tkinter.LabelFrame(productFrame,borderwidth=0,highlightthickness=0)
extraFrame.grid(row=3,column=0, sticky=W+E)
extra=ttk.Label(extraFrame,text="Additional")
extra.grid(row=0,column=0,padx=(30,10),pady=10, sticky="ew")
extraEntry=ttk.Entry(extraFrame,width=13)
extraEntry.grid(row=0,column=1,padx=30,sticky="w")


taxVar=tkinter.StringVar()
totVar=tkinter.StringVar()


#tax display
taxFrame=tkinter.LabelFrame(productFrame,borderwidth=0)
taxFrame.grid(row=3,column=1, sticky=W+E)
taxLabel=ttk.Label(taxFrame,text="Tax Amount")
taxLabel.grid(row=0,column=0,padx=(15,5),pady=10, sticky="w")
taxEntry=ttk.Entry(taxFrame,text=taxVar)
taxEntry.grid(row=0,column=1, sticky="w",padx=(0,15),)

tempTax=None
def subFunc():
    global tempTax,taxVar
    if extraEntry.get()!="":
        temp=float(subEntry.get())+float(extraEntry.get())
    else:
        temp=float(subEntry.get())
    tempTax=(temp*float(taxSelectEntry.get()[:-1]))/100
    taxVar.set(tempTax)
    totVar.set(temp+tempTax)
    totalEntry.focus_set()

extraEntry.bind('<Return>',lambda e: subFunc())
#tax selector
taxSelectFrame=tkinter.LabelFrame(productFrame,border=0,highlightthickness=0)
taxSelectFrame.grid(row=4,column=0,sticky=W+E)
taxTypeEntry=ttk.Combobox(taxSelectFrame,width=9,state="readonly",values=["CGST\SGST","IGST"])
taxTypeEntry.grid(row=4,column=0,padx=30,pady=10)
taxTypeEntry.current(0)
taxSelectEntry=ttk.Combobox(taxSelectFrame,width=3,state="readonly",values=["0%","5%","10%","12%","14%","18%","24%","28%"])
taxSelectEntry.grid(row=4,column=1,padx=0,pady=10)
taxSelectEntry.current(5)
    
#grand total
totalFrame=tkinter.LabelFrame(productFrame,border=0)
totalFrame.grid(row=4,column=1, sticky=W+E)
totalLabel=ttk.Label(totalFrame,text="Total")
totalLabel.grid(row=0,column=0,padx=30,pady=10, sticky="w")
totalEntry=ttk.Entry(totalFrame,text=totVar)
totalEntry.grid(row=0,column=1,sticky="w",padx=(5,1))
subEntry.bind('<Return>',lambda e: subFunc())

#button function
def enterFunc():
    enterFunction()

#cleaner function
def cleaner():
    invoiceNoEntry.delete(0, tkinter.END)
    gstNoEntry.delete(0, tkinter.END)
    bNameEntry.delete(0, tkinter.END)
    bAddEntry.delete(0, tkinter.END)
    productEntry.delete(0, tkinter.END)
    hsnEntry.delete(0, tkinter.END)
    subEntry.delete(0, tkinter.END)
    extraEntry.delete(0, tkinter.END)
    taxEntry.delete(0, tkinter.END)
    totalEntry.delete(0, tkinter.END)
    invoiceNoEntry.focus_set()
def reuse():
    enterFunc()
    invoiceNoEntry.delete(0, tkinter.END)
    productEntry.delete(0, tkinter.END)
    hsnEntry.delete(0, tkinter.END)
    subEntry.delete(0, tkinter.END)
    extraEntry.delete(0, tkinter.END)
    taxEntry.delete(0, tkinter.END)
    totalEntry.delete(0, tkinter.END)
    invoiceNoEntry.focus_set()

def submitCaller():
    enterFunc()
    cleaner()

#buttons
buttonFrame=tkinter.LabelFrame(productFrame,border=0,highlightthickness=0)
buttonFrame.grid(row=5,column=0, sticky=W+E)
clearButton=ttk.Button(buttonFrame,text="Clear",command=cleaner)
clearButton.grid(row=0,column=0,padx=15)
reuseButton=ttk.Button(buttonFrame,text="Submit & Reuse",command=reuse)
reuseButton.grid(row=0,column=1,padx=15)

enterButton=ttk.Button(productFrame,command=submitCaller,text="Submit")
enterButton.grid(row=5,column=1,sticky='ns',pady=20)
totalEntry.bind('<Return>',lambda e: enterButton.focus_set())
enterButton.bind('<Return>',lambda e: submitCaller())

#getting the values
def enterFunction():
    getInvoiceNo=invoiceNoEntry.get()
    getInvoiceDate=invoiceDateEntry.get()
    getGstNo=gstNoEntry.get()
    getName=bNameEntry.get()
    getAdd=bAddEntry.get()
    getProduct=productEntry.get()
    getHSN=hsnEntry.get()
    getSub=subEntry.get()
    getTaxType=taxTypeEntry.get()
    getTaxPercent=taxSelectEntry.get()
    getTotal=totalEntry.get()

    month=entryMonthEntry.get()
    year=entryYearEntry.get()
    bookpath=month+"_"+year

    path="D:\\fullsafe\\"+year+"\\"+month+"\\"
    Path=path+"Purchase_"+bookpath+".xlsx"

    if not os.path.exists(path):
        os.makedirs(path)
    if not os.path.exists(Path):
        openBook(Path)
    
    temp1,temp2,percent1,percent2=None,None,None,None
    if not extraEntry.get()=="":
        temp=float(subEntry.get())+float(extraEntry.get())
        igst=(temp*float(getTaxPercent[:-1]))/100
        gst=(temp*float(getTaxPercent[:-1]))/200
    else:
        igst=(float(getSub)*float(getTaxPercent[:-1]))/100
        gst=(float(getSub)*float(getTaxPercent[:-1]))/200
    tempGst=igst
    if getTaxType=="IGST":
        temp2="-"
        percent1=getTaxPercent
        percent2="-"
        gst="-"
    else:
        temp1="-"
        percent1="-"
        percent2=str(int(getTaxPercent[:-1])/2)+"%"
        igst="-"
    list=[getInvoiceNo,getInvoiceDate,getGstNo,getName,getAdd,getProduct,getHSN,getSub,percent1,igst,percent2,gst,percent2,gst,tempGst,float(getTotal)]
    book=openpyxl.load_workbook(Path)
    sheet=book.active
    sheet.append(list)
    book.save(Path)
chang()
sv_ttk.set_theme(themeSet) #sets theme
root.mainloop()
